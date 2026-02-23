#!/usr/bin/env python3
from __future__ import annotations

from pathlib import Path
import re

import pandas as pd

SOURCE_DIR = Path("E_report")
OUT_DIR = Path("data/processed")
MERGED_CSV = OUT_DIR / "all_merged.csv"
MERGED_XLSX = OUT_DIR / "all_merged.xlsx"
REPORT = OUT_DIR / "merge_report.md"

ID_HINTS = ["skole", "school", "school_name", "enhetsnummer", "orgnr", "unit", "id"]


def read_csv_flexible(path: Path) -> pd.DataFrame:
    """Read CSV files with flexible delimiter + encoding handling.

    UDIR exports and manually downloaded regional files can use different
    encodings (utf-8, utf-8-sig, cp1252, latin-1) and delimiters.
    """

    separators = [",", ";", "\t", "|"]
    encodings = ["utf-8", "utf-8-sig", "cp1252", "latin-1"]
    last_error: Exception | None = None

    for enc in encodings:
        for sep in separators:
            try:
                df = pd.read_csv(path, sep=sep, encoding=enc)
                if df.shape[1] > 1:
                    return df
            except Exception as exc:
                last_error = exc
                continue

    # Last attempt: let pandas auto-detect separator with python engine,
    # while still trying practical encodings.
    for enc in encodings:
        try:
            return pd.read_csv(path, sep=None, engine="python", encoding=enc)
        except Exception as exc:
            last_error = exc

    if last_error is not None:
        raise last_error

    # Defensive fallback (should never hit in practice)
    return pd.read_csv(path)


def _normalize_text(text: str) -> str:
    text = str(text).strip().lower()
    text = re.sub(r"\s+", "_", text)
    return re.sub(r"[^\w]+", "_", text).strip("_")


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    normalized: list[str] = []
    used: dict[str, int] = {}

    for col in df.columns:
        base = _normalize_text(str(col)) or "col"
        n = used.get(base, 0)
        used[base] = n + 1
        normalized.append(base if n == 0 else f"{base}_{n + 1}")

    df.columns = normalized
    return df


def pick_id_column(df: pd.DataFrame) -> str | None:
    for col in df.columns:
        if any(h in col.lower() for h in ID_HINTS):
            return col
    return None


def reorder_columns(df: pd.DataFrame) -> pd.DataFrame:
    priority_patterns = [
        "source_file",
        "skole",
        "school",
        "enhetsnummer",
        "orgnr",
        "kommune",
        "municipality",
        "trinn",
        "grade",
        "ar",
        "year",
    ]

    cols = list(df.columns)
    picked: list[str] = []

    for pattern in priority_patterns:
        for col in cols:
            if col in picked:
                continue
            if pattern in col.lower():
                picked.append(col)

    rest = [c for c in cols if c not in picked]
    rest = sorted(rest, key=lambda x: x.lower())

    return df[picked + rest]


def prettify_for_excel(path: Path, sheet_name: str = "merged") -> None:
    # openpyxl is optional; function is called only when available.
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = load_workbook(path)
    ws = wb[sheet_name]

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_cells in ws.columns:
        max_len = 0
        for c in col_cells:
            v = "" if c.value is None else str(c.value)
            max_len = max(max_len, len(v))
            if c.row > 1:
                c.alignment = Alignment(vertical="top", wrap_text=False)
        adjusted = min(max(max_len + 2, 10), 60)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = adjusted

    wb.save(path)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    SOURCE_DIR.mkdir(parents=True, exist_ok=True)

    files = sorted(SOURCE_DIR.glob("*.csv"))
    if not files:
        raise SystemExit(
            "No CSV files found in E_report/. Please place your downloaded files there, then rerun."
        )

    normalized_frames: list[tuple[str, pd.DataFrame, str | None]] = []
    notes: list[str] = []

    for csv_path in files:
        df = normalize_columns(read_csv_flexible(csv_path))
        id_col = pick_id_column(df)

        if id_col is None:
            notes.append(f"- `{csv_path.name}`: no clear unique ID column found.")
        elif "skole" not in id_col and "school" not in id_col:
            notes.append(f"- `{csv_path.name}`: merged by `{id_col}` (not school name).")

        df["source_file"] = csv_path.name
        normalized_frames.append((csv_path.name, df, id_col))

    merged: pd.DataFrame | None = None
    merge_key: str | None = None

    for _, df, id_col in normalized_frames:
        if merged is None:
            merged = df
            merge_key = id_col
            continue

        if merge_key and merge_key in df.columns:
            key = merge_key
        elif id_col and id_col in merged.columns and id_col in df.columns:
            key = id_col
            merge_key = key
        else:
            common = [c for c in merged.columns if c in df.columns and c != "source_file"]
            key = common[0] if common else None

        if key:
            merged = merged.merge(df, on=key, how="outer", suffixes=("", "_dup"))
        else:
            merged = pd.concat([merged, df], ignore_index=True, sort=False)

    assert merged is not None

    dup_cols = [c for c in merged.columns if c.endswith("_dup")]
    if dup_cols:
        merged = merged.drop(columns=dup_cols)

    merged = reorder_columns(merged)
    merged = merged.sort_values(by=["source_file"], kind="stable", na_position="last")

    merged.to_csv(MERGED_CSV, index=False, encoding="utf-8-sig")

    excel_created = False
    try:
        with pd.ExcelWriter(MERGED_XLSX, engine="openpyxl") as writer:
            merged.to_excel(writer, index=False, sheet_name="merged")
        prettify_for_excel(MERGED_XLSX)
        excel_created = True
    except ModuleNotFoundError:
        notes.append("- openpyxl not installed; skipped formatted Excel output.")

    REPORT.write_text(
        "# Merge report\n\n"
        f"- Source directory: `{SOURCE_DIR}`\n"
        f"- Total input files: **{len(files)}**\n"
        f"- Total output rows: **{len(merged)}**\n"
        f"- Total output columns: **{len(merged.columns)}**\n\n"
        "## Notes\n\n"
        + ("\n".join(notes) if notes else "- All files appear to have a school-related merge key.")
        + "\n",
        encoding="utf-8",
    )

    print(f"Merged CSV saved to: {MERGED_CSV}")
    if excel_created:
        print(f"Formatted Excel saved to: {MERGED_XLSX}")
    else:
        print("Formatted Excel skipped: openpyxl is not installed.")
    print(f"Report saved to: {REPORT}")


if __name__ == "__main__":
    main()
